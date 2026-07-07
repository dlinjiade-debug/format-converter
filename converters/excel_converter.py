"""Excel 和 CSV 相关格式转换"""

from pathlib import Path


def excel_to_pdf(input_path: Path, output_path: Path) -> Path:
    """Excel → PDF"""
    from openpyxl import load_workbook
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    font_name = "Helvetica"
    for fp in ["C:/Windows/Fonts/simhei.ttf", "C:/Windows/Fonts/simsun.ttc"]:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                font_name = "ChineseFont"
                break
            except Exception:
                continue

    wb = load_workbook(str(input_path))
    page_size = landscape(A4)
    pdf_doc = SimpleDocTemplate(str(output_path), pagesize=page_size)
    header_style = ParagraphStyle("Header", fontName=font_name, fontSize=9, leading=12)
    cell_style = ParagraphStyle("Cell", fontName=font_name, fontSize=8, leading=11)
    title_style = ParagraphStyle("Title", fontName=font_name, fontSize=14, leading=20, spaceAfter=12)

    elements = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if elements:
            elements.append(PageBreak())
        elements.append(Paragraph(f"工作表: {sheet_name}", title_style))

        rows = []
        for row in ws.iter_rows(values_only=True):
            row_data = []
            for cell in row:
                val = str(cell) if cell is not None else ""
                row_data.append(Paragraph(val.replace("&", "&amp;").replace("<", "&lt;"), cell_style))
            if any(str(c) != "" for c in row if c is not None):
                rows.append(row_data)

        if rows:
            # 限制列宽
            col_count = max(len(r) for r in rows) if rows else 1
            available_width = page_size[0] - 100
            col_width = min(available_width / col_count, 150)

            table = Table(rows, colWidths=[col_width] * col_count)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]
            table.setStyle(TableStyle(style_cmds))
            elements.append(table)
        else:
            elements.append(Paragraph("（空工作表）", cell_style))

    if not elements:
        elements.append(Spacer(1, 1))

    pdf_doc.build(elements)
    return output_path


def excel_to_csv(input_path: Path, output_path: Path) -> Path:
    """Excel → CSV"""
    from openpyxl import load_workbook
    import csv

    wb = load_workbook(str(input_path), data_only=True)
    ws = wb.active

    with open(str(output_path), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(cell) if cell is not None else "" for cell in row])

    return output_path


def csv_to_excel(input_path: Path, output_path: Path) -> Path:
    """CSV → Excel"""
    from openpyxl import Workbook
    import csv

    wb = Workbook()
    ws = wb.active

    with open(str(input_path), "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    wb.save(str(output_path))
    return output_path
